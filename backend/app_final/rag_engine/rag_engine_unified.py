"""
Rewritten RAG engine with:
- Empty-query behavior: when query is blank, skip FAISS and return rows filtered via SQLite.
- Stable embedding pipeline with SentenceTransformer (normalized for cosine via Inner Product).
- FAISS artifacts persisted on disk (index, id_map, hashes).
- Chunked DB streaming for (re)builds.
- CSV utilities with clean, user-friendly formatting and readable column headers.
- Filter application supporting city, permit_type, class, work_class, status, and date ranges.

Author: ChatGPT
Modified: Enhanced CSV formatting for better user experience
"""

from __future__ import annotations
import logging
import os
import io
import csv
import json
import time
import faiss
import sqlite3
import hashlib
import numpy as np
import re
from typing import List, Dict, Any, Optional, Iterable, Tuple
from sentence_transformers import SentenceTransformer

logger = logging.getLogger(__name__)


# ----------------------------- Helpers -----------------------------
def _safe(s: Any) -> str:
    return "" if s is None else str(s)


def _row_to_text(row: Dict[str, Any]) -> str:
    """DESCRIPTION-ONLY index for better semantic search"""
    description = _safe(row.get('description'))
    if not description or description.strip() == "":
        return "no description available"
    return f"project: {description}"  # Focus only on description


def _row_to_description_only(row: Dict[str, Any]) -> str:
    """
    Build text representation using ONLY the description field for keyword-based search.
    This is used when you want to search only in permit descriptions.
    """
    description = _safe(row.get('description'))
    if not description or description.strip() == "":
        return "no description available"
    return description.strip()


def _row_to_text_keyword_search(row: Dict[str, Any]) -> str:
    """
    Build text representation optimized for keyword search in description field.
    Includes minimal metadata but focuses on description for semantic search.
    """
    description = _safe(row.get('description'))
    if not description or description.strip() == "":
        description = "no description available"

    # Include minimal metadata for context but focus on description
    parts = [
        f"description: {description}",
        f"permit_type: {_safe(row.get('permit_type'))}",
        f"work_class: {_safe(row.get('work_class'))}",
    ]
    return " | ".join(parts).strip()


# ----------------------------- RAGIndex -----------------------------
class RAGIndex:
    """
    Persistent RAG index over the SQLite permits table.

    Artifacts (in index_dir):

      - index.faiss        : FAISS IndexFlatIP with normalized vectors
      - id_map.npy         : numpy int64 array mapping FAISS row -> permit_id
      - hashes.json        : map permit_id -> md5(text_recipe) (for future incremental)
    """

    def __init__(
            self,
            db_path: str,
            index_dir: str = "rag_index",
            model_name: str = "all-MiniLM-L6-v2",
    ) -> None:
        self.db_path = db_path
        self.index_dir = index_dir
        self.index_path = os.path.join(index_dir, "index.faiss")
        self.idmap_path = os.path.join(index_dir, "id_map.npy")
        self.hashes_path = os.path.join(index_dir, "hashes.json")
        self.model_name = model_name

        os.makedirs(self.index_dir, exist_ok=True)

        self._model: Optional[SentenceTransformer] = None
        self.index: Optional[faiss.Index] = None
        self.id_map: Optional[np.ndarray] = None  # numpy array of permit_ids (int64)

    # ---------- Model ----------
    @property
    def model(self) -> SentenceTransformer:
        if self._model is None:
            # normalize_embeddings=True => cosine similarity via Inner Product
            self._model = SentenceTransformer(self.model_name)
        return self._model

    def embedding_dim(self) -> int:
        try:
            return int(self.model.get_sentence_embedding_dimension())
        except Exception:
            # Fallback to common dim for MiniLM
            return 384

    def _encode(self, texts: List[str], batch_size: int = 256) -> np.ndarray:
        # sentence-transformers can normalize for us (v2+)
        return self.model.encode(
            texts,
            batch_size=batch_size,
            normalize_embeddings=True,
            convert_to_numpy=True,
            show_progress_bar=False,
        )

    # ---------- DB ----------
    def _connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.db_path)

    def _fetch_permits_iter(self, chunk_size: int = 2000) -> Iterable[List[Dict[str, Any]]]:
        """
        Stream rows from DB in chunks to avoid loading entire table in memory.
        """
        cols = [
            "id",
            "city",
            "permit_num",
            "permit_type",
            "permit_class_mapped",
            "work_class",
            "description",
            "applied_date",
            "issued_date",
            "current_status",
            "applicant_name",
            "applicant_address",
            "contractor_name",
            "contractor_address",
            "contractor_company_name",
            "contractor_phone",
        ]
        conn = self._connect()
        try:
            cur = conn.cursor()
            cur.execute(f"SELECT {', '.join(cols)} FROM permits")
            while True:
                rows = cur.fetchmany(chunk_size)
                if not rows:
                    break
                batch: List[Dict[str, Any]] = []
                for r in rows:
                    d = dict(zip(cols, r))
                    batch.append(d)
                yield batch
        finally:
            conn.close()

    def search_heater_test(self, query: str, filters: Dict[str, Any], top_k: int = 10) -> List[Dict[str, Any]]:
        """TEMPORARY TEST METHOD: Search heaters with large historical scope"""

        print(f"🧪 HEATER TEST SEARCH: '{query}'")
        print(f"   🔧 Filters: {filters}")

        # Get filtered permits with MUCH larger limit
        conn = self._connect()
        try:
            sql_parts = ["SELECT * FROM permits WHERE 1=1"]
            params = []

            # Apply filters (simplified)
            if filters.get("city"):
                cities = filters["city"]
                if isinstance(cities, list):
                    city_conditions = []
                    for city in cities:
                        city_conditions.append("LOWER(TRIM(city)) = LOWER(TRIM(?))")
                        params.append(city.strip())
                    sql_parts.append(f"AND ({' OR '.join(city_conditions)})")

            if filters.get("permit_type"):
                ptypes = filters["permit_type"]
                if isinstance(ptypes, list):
                    type_conditions = []
                    for ptype in ptypes:
                        type_conditions.append("LOWER(TRIM(permit_type)) = LOWER(TRIM(?))")
                        params.append(ptype.strip())
                    sql_parts.append(f"AND ({' OR '.join(type_conditions)})")

            if filters.get("permit_class_mapped"):
                classes = filters["permit_class_mapped"]
                if isinstance(classes, list):
                    class_conditions = []
                    for pclass in classes:
                        class_conditions.append("LOWER(TRIM(permit_class_mapped)) = LOWER(TRIM(?))")
                        params.append(pclass.strip())
                    sql_parts.append(f"AND ({' OR '.join(class_conditions)})")

            if filters.get("work_class"):
                work_classes = filters["work_class"]
                if isinstance(work_classes, list):
                    work_conditions = []
                    for work_class in work_classes:
                        work_conditions.append("LOWER(TRIM(work_class)) = LOWER(TRIM(?))")
                        params.append(work_class.strip())
                    sql_parts.append(f"AND ({' OR '.join(work_conditions)})")

            # FIXED: Use random ordering and large limit to include historical data
            sql_parts.append("ORDER BY RANDOM() LIMIT ?")
            params.append(2000)  # Much larger limit

            sql = " ".join(sql_parts)
            print(f"   🗄️ SQL: {sql}")
            print(f"   📝 Params: {params}")

            cur = conn.cursor()
            cur.execute(sql, params)

            columns = [desc[0] for desc in cur.description]
            filtered_permits = [dict(zip(columns, row)) for row in cur.fetchall()]

            print(f"   📊 Database returned: {len(filtered_permits)} permits")

            if filtered_permits:
                # Show ID range
                min_id = min(int(p['id']) for p in filtered_permits)
                max_id = max(int(p['id']) for p in filtered_permits)
                print(f"   🆔 ID Range: {min_id} to {max_id}")

                # Look for heater permits specifically
                heater_permits = [p for p in filtered_permits if 'heater' in str(p.get('description', '')).lower()]
                print(f"   🔥 Found {len(heater_permits)} permits with 'heater' in description")

                if heater_permits:
                    sample_heater = heater_permits[0]
                    print(
                        f"   🔥 Sample heater permit: ID={sample_heater['id']}, Desc: {sample_heater.get('description', '')[:100]}...")

            conn.close()

            # Apply semantic search if available
            print(f"🧠 SEMANTIC SEARCH STEP:")
            print(query)
            print("*********")
            print(self.index)
            print("*********")
            print(query.strip())
            print("*********")
            print(self.index)
            print("&&&&&&&&&&")
            print(self.id_map)
            if query and query.strip() and self.index is not None and self.id_map is not None:
                print(f"   🧠 Applying semantic search...")
                results = self._semantic_search_within_permits(filtered_permits, query, top_k, True)
                print(f"   🎯 Semantic results: {len(results)}")
            else:
                print(f"   📋 No semantic search, returning first {top_k} results")
                results = filtered_permits[:top_k]
                for permit in results:
                    permit['_rag_score'] = 1.0

            return results

        except Exception as e:
            print(f"   ❌ Error: {e}")
            if conn:
                conn.close()
            return []

    def search_fixed_debug(self, query: str, top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
                           oversample: int = 5, max_oversample: int = 80, return_scores: bool = False) -> List[
        Dict[str, Any]]:
        """FIXED VERSION: Enhanced search with proper database limits"""

        logger.info(f"🔍 DEBUG SEARCH STARTING:")
        logger.info(f"   📝 Query: '{query}'")
        logger.info(f"   🔧 Filters: {filters}")

        try:
            # Verify index is loaded
            if self.index is None or self.id_map is None:
                if not self.load():
                    logger.error(f"   ❌ Failed to load index from disk")
                    return []

            # FIXED: Use larger database limits to include historical data
            if filters and any(filters.values()):
                db_limit = max(top_k * oversample, 1000)  # At least 1000 permits
                filtered_permits = self._get_filtered_permits_from_db_debug(filters, db_limit)
            else:
                db_limit = max(top_k * 3, 500)  # At least 500 permits
                filtered_permits = self._get_recent_permits_simple(db_limit)

            if not filtered_permits:
                logger.warning(f"   ⚠️ No permits after database filtering")
                return []

            # Apply semantic search if query provided
            if query and query.strip():
                results = self._semantic_search_within_permits_debug(filtered_permits, query, top_k, return_scores)
            else:
                results = filtered_permits[:top_k]
                if return_scores:
                    for permit in results:
                        permit['_rag_score'] = 1.0

            logger.info(f"✅ SEARCH COMPLETE: {len(results)} results")
            return results

        except Exception as e:
            logger.error(f"❌ SEARCH ERROR: {e}")
            return []

    def _get_filtered_permits_from_db_debug(self, filters: Dict[str, Any], limit: int) -> List[Dict[str, Any]]:
        """Enhanced database filtering with detailed debugging - FIXED ORDERING"""

        logger.info(f"   🗄️ DATABASE FILTER DEBUG:")
        logger.info(f"      📝 Input filters: {filters}")
        logger.info(f"      📊 Limit: {limit}")

        conn = self._connect()
        try:
            sql_parts = ["SELECT * FROM permits WHERE 1=1"]
            params = []
            applied_filters = []

            # City filter with debugging
            if filters.get("city"):
                cities = filters["city"]
                if isinstance(cities, list) and cities:
                    city_conditions = []
                    for city in cities:
                        city_conditions.append("LOWER(TRIM(city)) = LOWER(TRIM(?))")
                        params.append(city.strip())
                    sql_parts.append(f"AND ({' OR '.join(city_conditions)})")
                    applied_filters.append(f"city IN {cities}")
                    logger.info(f"      📍 City filter: {cities}")

            # Permit type filter with debugging
            if filters.get("permit_type"):
                ptypes = filters["permit_type"]
                if isinstance(ptypes, list) and ptypes:
                    type_conditions = []
                    for ptype in ptypes:
                        type_conditions.append("LOWER(TRIM(permit_type)) = LOWER(TRIM(?))")
                        params.append(ptype.strip())
                    sql_parts.append(f"AND ({' OR '.join(type_conditions)})")
                    applied_filters.append(f"permit_type IN {ptypes}")
                    logger.info(f"      🏗️ Permit type filter: {ptypes}")

            # Permit class filter with debugging
            if filters.get("permit_class_mapped"):
                classes = filters["permit_class_mapped"]
                if isinstance(classes, list) and classes:
                    class_conditions = []
                    for pclass in classes:
                        class_conditions.append("LOWER(TRIM(permit_class_mapped)) = LOWER(TRIM(?))")
                        params.append(pclass.strip())
                    sql_parts.append(f"AND ({' OR '.join(class_conditions)})")
                    applied_filters.append(f"permit_class_mapped IN {classes}")
                    logger.info(f"      🏷️ Permit class filter: {classes}")

            # Work class filter with debugging - FIXED CASE SENSITIVITY
            if filters.get("work_class"):
                work_classes = filters["work_class"]
                if isinstance(work_classes, list) and work_classes:
                    work_conditions = []
                    for work_class in work_classes:
                        work_conditions.append("LOWER(TRIM(work_class)) = LOWER(TRIM(?))")
                        params.append(work_class.strip())
                    sql_parts.append(f"AND ({' OR '.join(work_conditions)})")
                    applied_filters.append(f"work_class IN {work_classes}")
                    logger.info(f"      ⚒️ Work class filter: {work_classes}")

            # FIXED: Change ordering to include historical data
            # Instead of "ORDER BY issued_date DESC LIMIT ?" which only gets recent permits
            # Use mixed ordering to get both recent and historical data
            if limit > 500:
                # For large limits, use random ordering to get diverse results
                sql_parts.append("ORDER BY RANDOM() LIMIT ?")
                logger.info(f"      🔀 Using RANDOM() ordering for diverse results (limit: {limit})")
            else:
                # For smaller limits, still prioritize recent but include more
                sql_parts.append("ORDER BY issued_date DESC LIMIT ?")
                logger.info(f"      📅 Using date ordering for recent results (limit: {limit})")

            params.append(limit)

            # Execute query with full logging
            sql = " ".join(sql_parts)
            logger.info(f"      🗄️ Final SQL: {sql}")
            logger.info(f"      📝 Parameters: {params}")
            logger.info(f"      🔧 Applied filters: {applied_filters}")

            cur = conn.cursor()
            start_time = time.time()
            cur.execute(sql, params)
            query_time = time.time() - start_time

            # Convert to dict format
            columns = [desc[0] for desc in cur.description]
            results = [dict(zip(columns, row)) for row in cur.fetchall()]

            logger.info(f"      ⏱️ Query execution time: {query_time:.3f}s")
            logger.info(f"      ✅ Database returned: {len(results)} permits")

            # Log sample results for verification INCLUDING ID RANGE
            if results:
                sample = results[0]
                min_id = min(int(r['id']) for r in results)
                max_id = max(int(r['id']) for r in results)
                logger.info(f"      📋 Sample result: ID={sample.get('id')}, "
                            f"city={sample.get('city')}, "
                            f"permit_type={sample.get('permit_type')}, "
                            f"work_class={sample.get('work_class')}")
                logger.info(f"      🆔 ID Range: {min_id} to {max_id}")

            return results

        except Exception as e:
            logger.error(f"      ❌ Database filter error: {e}")
            return []
        finally:
            conn.close()

    def search_with_historical_fallback(self, query: str, top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
                                        oversample: int = 5, return_scores: bool = False) -> List[Dict[str, Any]]:
        """Search with automatic fallback to historical data if no recent results"""

        logger.info(f"🔍 HISTORICAL FALLBACK SEARCH: '{query}'")

        # Try with recent data first (smaller limit)
        recent_results = self.search_fixed_debug(
            query=query,
            top_k=top_k,
            filters=filters,
            oversample=oversample,
            return_scores=return_scores
        )

        if len(recent_results) >= top_k // 2:  # If we got at least half the results we wanted
            logger.info(f"   ✅ Found {len(recent_results)} recent results, no fallback needed")
            return recent_results

        # Not enough recent results, search with much larger historical scope
        logger.info(f"   ⚠️ Only {len(recent_results)} recent results, searching historical data...")

        # Force larger database limit for historical search
        if filters and any(filters.values()):
            filtered_permits = self._get_filtered_permits_from_db_debug(filters, 2000)  # Much larger limit
        else:
            filtered_permits = self._get_recent_permits_simple(2000)

        logger.info(f"   📊 Historical search pool: {len(filtered_permits)} permits")

        if query and query.strip():
            historical_results = self._semantic_search_within_permits_debug(
                filtered_permits, query, top_k, return_scores
            )
        else:
            historical_results = filtered_permits[:top_k]
            if return_scores:
                for permit in historical_results:
                    permit['_rag_score'] = 1.0

        logger.info(f"   📊 Historical search results: {len(historical_results)}")
        return historical_results

    def _semantic_search_within_permits_debug(self, permits: List[Dict[str, Any]], query: str, top_k: int,
                                              return_scores: bool) -> List[Dict[str, Any]]:
        """Enhanced semantic search with comprehensive debugging"""

        logger.info(f"   🧠 SEMANTIC SEARCH DEBUG:")
        logger.info(f"      📊 Input permits: {len(permits)}")
        logger.info(f"      🔎 Query: '{query}'")
        logger.info(f"      🎯 Target results: {top_k}")

        if not permits:
            logger.warning(f"      ⚠️ No permits to search within")
            return []

        try:
            # Get permit IDs from filtered results
            permit_ids = set(int(p['id']) for p in permits)
            logger.info(f"      🆔 Filtered permit IDs count: {len(permit_ids)}")
            logger.info(f"      🆔 Sample filtered IDs: {sorted(list(permit_ids))[:10]}")

            # Check FAISS index status
            if self.index is None or self.id_map is None:
                logger.error(f"      ❌ FAISS index not available")
                return self._simple_text_search_debug(permits, query, top_k, return_scores)

            # Debug FAISS index contents
            faiss_ids = set(self.id_map.tolist())
            logger.info(f"      🗂️ FAISS index contains: {len(faiss_ids)} permit IDs")
            logger.info(f"      🗂️ Sample FAISS IDs: {sorted(list(faiss_ids))[:10]}")

            # Check overlap between filtered permits and FAISS index
            overlap_ids = permit_ids.intersection(faiss_ids)
            logger.info(f"      🔗 Overlap between filtered and FAISS: {len(overlap_ids)} permits")
            logger.info(f"      🔗 Sample overlap IDs: {sorted(list(overlap_ids))[:10]}")

            if len(overlap_ids) == 0:
                logger.error(f"      ❌ NO OVERLAP! Filtered permits not in FAISS index")
                logger.error(f"      ❌ This suggests FAISS index is out of sync with database")
                logger.error(f"      🔄 Falling back to simple text search")
                return self._simple_text_search_debug(permits, query, top_k, return_scores)

            overlap_percentage = (len(overlap_ids) / len(permit_ids)) * 100
            logger.info(f"      📊 Overlap percentage: {overlap_percentage:.1f}%")

            if overlap_percentage < 50:
                logger.warning(f"      ⚠️ Low overlap ({overlap_percentage:.1f}%) - FAISS index may be outdated")

            # Create query embedding
            logger.info(f"      🧮 Creating query embedding...")
            start_time = time.time()
            query_embedding = self._encode([query.strip()])[0].reshape(1, -1)
            embed_time = time.time() - start_time
            logger.info(f"      ✅ Query embedding created in {embed_time:.3f}s, shape: {query_embedding.shape}")

            # Search FAISS index
            # search_count = min(len(self.id_map), max(top_k * 5, 100))
            search_count = min(len(self.id_map), 50000)
            logger.info(f"      🔍 Searching FAISS for top {search_count} results...")

            start_time = time.time()
            scores, indices = self.index.search(query_embedding, search_count)
            search_time = time.time() - start_time
            logger.info(f"      ✅ FAISS search completed in {search_time:.3f}s")
            logger.info(f"      📊 FAISS returned {len(indices[0])} candidate results")

            # Match FAISS results with filtered permits
            results = []
            matches_found = 0
            total_candidates = 0

            for idx, score in zip(indices[0], scores[0]):
                total_candidates += 1
                if idx >= 0:  # Valid FAISS index
                    permit_id = int(self.id_map[idx])

                    # Only include if this permit was in our filtered results
                    if permit_id in permit_ids:
                        matches_found += 1
                        # Find the full permit data from filtered permits
                        permit_data = next((p for p in permits if int(p['id']) == permit_id), None)
                        if permit_data:
                            if return_scores:
                                permit_data['_rag_score'] = float(score)
                            results.append(permit_data)

                            logger.info(f"         ✅ Match {matches_found}: Permit {permit_id}, Score: {score:.4f}")

                            if len(results) >= top_k:
                                break

            logger.info(f"      📊 Search summary:")
            logger.info(f"         🔍 Total FAISS candidates: {total_candidates}")
            logger.info(f"         🎯 Matches in filtered set: {matches_found}")
            logger.info(f"         ✅ Final results: {len(results)}")

            if len(results) == 0:
                logger.warning(f"      ⚠️ No semantic matches found, falling back to text search")
                return self._simple_text_search_debug(permits, query, top_k, return_scores)

            return results

        except Exception as e:
            logger.error(f"      ❌ Semantic search error: {e}")
            logger.error(f"      ↳ Falling back to simple text search")
            return self._simple_text_search_debug(permits, query, top_k, return_scores)

    def _simple_text_search_debug(self, permits: List[Dict[str, Any]], query: str, top_k: int, return_scores: bool) -> \
            List[Dict[str, Any]]:
        """Enhanced text search fallback with debugging"""

        logger.info(f"   📝 TEXT SEARCH FALLBACK:")
        logger.info(f"      📊 Searching {len(permits)} permits for: '{query}'")

        query_lower = query.lower().strip()
        scored_permits = []

        for permit in permits:
            description = str(permit.get('description', '')).lower()
            if query_lower in description:
                # Enhanced scoring
                score = 0
                score += description.count(query_lower) * 10  # Frequency

                # Bonus for word boundaries
                words = description.split()
                if query_lower in words:
                    score += 20

                # Bonus for early position
                pos = description.find(query_lower)
                if pos < 50:  # Found in first 50 characters
                    score += 10

                if return_scores:
                    permit['_rag_score'] = score
                scored_permits.append(permit)

        # Sort by score if available
        if return_scores and scored_permits:
            scored_permits.sort(key=lambda x: x.get('_rag_score', 0), reverse=True)

        result = scored_permits[:top_k]
        logger.info(f"      ✅ Text search found: {len(result)} matches")

        if result and return_scores:
            top_scores = [r.get('_rag_score', 0) for r in result[:3]]
            logger.info(f"      🎯 Top 3 text scores: {top_scores}")

        return result

    # ============================================================================
    # STEP 2: Force Rebuild Index Method
    # ============================================================================

    def force_rebuild_index_debug(self):
        """Force complete rebuild with detailed logging"""
        try:
            logger.info("🔄 FORCE INDEX REBUILD STARTING:")

            # Check current index status
            current_status = self.status()
            logger.info(f"   📊 Current index status: {current_status}")

            # Check database
            conn = self._connect()
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM permits")
            total_permits = cur.fetchone()[0]
            cur.execute("SELECT MIN(id), MAX(id) FROM permits")
            min_id, max_id = cur.fetchone()
            conn.close()

            logger.info(f"   🗄️ Database stats:")
            logger.info(f"      📊 Total permits: {total_permits}")
            logger.info(f"      🆔 ID range: {min_id} to {max_id}")

            # Delete existing index files
            import os
            files_deleted = []
            for file_path, name in [
                (self.index_path, "index.faiss"),
                (self.idmap_path, "id_map.npy"),
                (self.hashes_path, "hashes.json")
            ]:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    files_deleted.append(name)
                    logger.info(f"   🗑️ Deleted: {name}")

            if files_deleted:
                logger.info(f"   🗑️ Cleaned up {len(files_deleted)} existing index files")
            else:
                logger.info(f"   📝 No existing index files to clean up")

            # Reset index in memory
            self.index = None
            self.id_map = None
            logger.info(f"   🧠 Reset in-memory index objects")

            # Build fresh index
            logger.info(f"   🔨 Starting fresh index build...")
            start_time = time.time()
            result = self.build(full_reindex=True, batch_size=256)
            build_time = time.time() - start_time

            logger.info(f"   ✅ Index build completed in {build_time:.2f}s")
            logger.info(f"   📊 Build result: {result}")

            # Verify new index
            new_status = self.status()
            logger.info(f"   📊 New index status: {new_status}")

            if new_status.get('loaded') and new_status.get('vectors', 0) > 0:
                logger.info(f"   ✅ INDEX REBUILD SUCCESS!")
                logger.info(f"   📊 {new_status.get('vectors')} permits indexed")
            else:
                logger.error(f"   ❌ INDEX REBUILD FAILED!")

            return result

        except Exception as e:
            logger.error(f"   ❌ REBUILD ERROR: {e}")
            raise e

    def _fetch_rows_by_ids(self, ids: List[int]) -> List[Dict[str, Any]]:
        if not ids:
            return []
        conn = self._connect()
        conn.row_factory = sqlite3.Row
        try:
            qmarks = ",".join("?" for _ in ids)
            cur = conn.cursor()
            cur.execute(f"SELECT * FROM permits WHERE id IN ({qmarks})", ids)
            return [dict(row) for row in cur.fetchall()]
        finally:
            conn.close()

    def _fetch_all_rows(self) -> List[Dict[str, Any]]:
        conn = self._connect()
        conn.row_factory = sqlite3.Row
        try:
            cur = conn.cursor()
            cur.execute("SELECT * FROM permits")
            return [dict(r) for r in cur.fetchall()]
        finally:
            conn.close()

    def _get_filtered_permits_from_db(self, filters: Optional[Dict[str, Any]], limit: int = 1000) -> List[
        Dict[str, Any]]:
        """Apply database filters first - FIXED to handle list format"""

        if not filters:
            # No filters, get recent permits
            conn = self._connect()
            try:
                cur = conn.cursor()
                cur.execute("SELECT * FROM permits ORDER BY issued_date DESC LIMIT ?", (limit,))
                columns = [desc[0] for desc in cur.description]
                results = [dict(zip(columns, row)) for row in cur.fetchall()]
                logger.info(f"   🗄️ No filters: returning {len(results)} recent permits")
                return results
            finally:
                conn.close()

        conn = self._connect()
        try:
            sql_parts = ["SELECT * FROM permits WHERE 1=1"]
            params = []

            # City filter - handle list format
            if filters.get("city"):
                cities = filters["city"]
                if isinstance(cities, list) and cities:
                    placeholders = ",".join("?" * len(cities))
                    sql_parts.append(f"AND LOWER(TRIM(city)) IN ({placeholders})")
                    params.extend([city.lower().strip() for city in cities])
                    logger.info(f"   📍 City filter: {cities}")

            # Permit type filter - handle list format
            if filters.get("permit_type"):
                ptypes = filters["permit_type"]
                if isinstance(ptypes, list) and ptypes:
                    placeholders = ",".join("?" * len(ptypes))
                    sql_parts.append(f"AND LOWER(TRIM(permit_type)) IN ({placeholders})")
                    params.extend([ptype.lower().strip() for ptype in ptypes])
                    logger.info(f"   🏗️ Permit type filter: {ptypes}")

            # Permit class filter - handle list format
            if filters.get("permit_class_mapped"):
                classes = filters["permit_class_mapped"]
                if isinstance(classes, list) and classes:
                    placeholders = ",".join("?" * len(classes))
                    sql_parts.append(f"AND LOWER(TRIM(permit_class_mapped)) IN ({placeholders})")
                    params.extend([cls.lower().strip() for cls in classes])
                    logger.info(f"   🏷️ Permit class filter: {classes}")

            # Work class filter - handle list format
            if filters.get("work_class"):
                work_classes = filters["work_class"]
                if isinstance(work_classes, list) and work_classes:
                    placeholders = ",".join("?" * len(work_classes))
                    sql_parts.append(f"AND LOWER(TRIM(work_class)) IN ({placeholders})")
                    params.extend([wc.lower().strip() for wc in work_classes])
                    logger.info(f"   ⚒️ Work class filter: {work_classes}")

            # Date filters
            if filters.get("issued_date_from"):
                sql_parts.append("AND issued_date >= ?")
                params.append(filters["issued_date_from"])
                logger.info(f"   📅 Date from: {filters['issued_date_from']}")

            if filters.get("issued_date_to"):
                sql_parts.append("AND issued_date <= ?")
                params.append(filters["issued_date_to"])
                logger.info(f"   📅 Date to: {filters['issued_date_to']}")

            # Add ordering and limit
            sql_parts.append("ORDER BY issued_date DESC LIMIT ?")
            params.append(limit)

            # Execute query
            sql = " ".join(sql_parts)
            logger.info(f"   🗄️ SQL: {sql}")
            logger.info(f"   📝 Params: {params}")

            cur = conn.cursor()
            cur.execute(sql, params)

            # Convert to dict format
            columns = [desc[0] for desc in cur.description]
            results = [dict(zip(columns, row)) for row in cur.fetchall()]

            logger.info(f"   ✅ Database filter result: {len(results)} permits")
            return results

        finally:
            conn.close()

    # ---------- Build / Save / Load ----------
    def build(self, full_reindex: bool = True, batch_size: int = 256) -> Dict[str, Any]:
        """
        Full rebuild of FAISS + ID map. (Incremental can be added later using hashes.)
        """
        start = time.time()
        all_texts: List[str] = []
        all_ids: List[int] = []
        hashes: Dict[int, str] = {}

        for chunk in self._fetch_permits_iter():
            for row in chunk:
                pid = int(row["id"])
                text = _row_to_text(row)
                h = hashlib.md5(text.encode("utf-8")).hexdigest()
                hashes[pid] = h
                all_texts.append(text)
                all_ids.append(pid)

        if not all_texts:
            # empty DB
            dim = self.embedding_dim()
            self.index = faiss.IndexFlatIP(dim)
            self.id_map = np.zeros((0,), dtype=np.int64)
            self._save_artifacts(hashes, start)
            return {"built": 0, "dim": dim, "took_s": round(time.time() - start, 2)}

        embs = self._encode(all_texts, batch_size=batch_size)  # normalized -> cosine via IP
        dim = embs.shape[1]

        # Build FlatIP index
        idx = faiss.IndexFlatIP(dim)
        idx.add(embs)

        self.index = idx
        self.id_map = np.array(all_ids, dtype=np.int64)

        self._save_artifacts(hashes, start)
        return {"built": len(all_ids), "dim": dim, "took_s": round(time.time() - start, 2)}

    def build_incremental(self, permit_ids: List[int], batch_size: int = 256) -> Dict[str, Any]:
        """
        Incremental build - only add new permits to existing index.

        Args:
            permit_ids: List of permit IDs to add to the index
            batch_size: Batch size for encoding
        """
        start = time.time()

        if not permit_ids:
            return {"built": 0, "message": "No permit IDs provided"}

        # Load existing index and hashes
        if not self.load():
            return {"error": "Cannot load existing index for incremental build"}

        # Load existing hashes to check what's already indexed
        existing_hashes = {}
        if os.path.exists(self.hashes_path):
            try:
                with open(self.hashes_path, "r", encoding="utf-8") as f:
                    existing_hashes = json.load(f)
            except Exception:
                existing_hashes = {}

        # Filter out permits that are already indexed
        new_permit_ids = []
        for pid in permit_ids:
            if str(pid) not in existing_hashes:
                new_permit_ids.append(pid)

        if not new_permit_ids:
            return {"built": 0, "dim": self.embedding_dim(), "took_s": round(time.time() - start, 2),
                    "message": "No new permits to index"}

        # Fetch only the new permits
        new_rows = self._fetch_rows_by_ids(new_permit_ids)

        all_texts: List[str] = []
        all_ids: List[int] = []
        hashes: Dict[int, str] = {}

        for row in new_rows:
            pid = int(row["id"])
            text = _row_to_text(row)
            h = hashlib.md5(text.encode("utf-8")).hexdigest()
            hashes[pid] = h
            all_texts.append(text)
            all_ids.append(pid)

        if not all_texts:
            return {"built": 0, "dim": self.embedding_dim(), "took_s": round(time.time() - start, 2),
                    "message": "No valid permits to index"}

        # Encode new texts
        new_embs = self._encode(all_texts, batch_size=batch_size)

        # Add new vectors to existing index
        self.index.add(new_embs)

        # Extend ID map
        self.id_map = np.concatenate([self.id_map, np.array(all_ids, dtype=np.int64)])

        # Merge hashes
        hashes.update(existing_hashes)

        # Save updated artifacts
        self._save_artifacts(hashes, start)

        return {
            "built": len(all_ids),
            "dim": self.embedding_dim(),
            "took_s": round(time.time() - start, 2),
            "type": "incremental",
            "new_permits": len(all_ids)
        }

    def _save_artifacts(self, hashes: Dict[int, str], start_time: float) -> None:
        if self.index is None or self.id_map is None:
            # Ensure on-disk files are at least consistent
            dim = self.embedding_dim()
            self.index = self.index or faiss.IndexFlatIP(dim)
            self.id_map = self.id_map or np.zeros((0,), dtype=np.int64)
        faiss.write_index(self.index, self.index_path)
        np.save(self.idmap_path, self.id_map)
        with open(self.hashes_path, "w", encoding="utf-8") as f:
            json.dump(hashes, f)
        # reload to be safe
        self.load()

    def load(self) -> bool:
        if not (os.path.exists(self.index_path) and os.path.exists(self.idmap_path)):
            return False
        self.index = faiss.read_index(self.index_path)
        self.id_map = np.load(self.idmap_path)
        return True

    def status(self) -> Dict[str, Any]:
        ok = self.index is not None and self.id_map is not None
        return {
            "loaded": ok,
            "index_path": self.index_path,
            "vectors": int(self.id_map.shape[0]) if self.id_map is not None else 0,
            "dim": int(self.index.d) if self.index is not None else None,
        }

    # ---------- Data Cleaning Helper Methods ----------
    def _clean_text(self, text: Any) -> str:
        """Clean and format text fields."""
        if text is None:
            return ""

        text = str(text).strip()

        # Convert common abbreviations to readable format
        replacements = {
            'INST': 'Install',
            'INSP': 'Inspection',
            'ELEC': 'Electrical',
            'MECH': 'Mechanical',
            'PLMB': 'Plumbing',
            'BLDG': 'Building',
            'RESI': 'Residential',
            'COMM': 'Commercial',
            'HVAC': 'HVAC',
            'DEMO': 'Demolition',
            'REMD': 'Remodel',
            'ADDN': 'Addition',
            'REP': 'Repair',
            'NEW': 'New'
        }

        # Apply replacements for whole words only
        words = text.split()
        cleaned_words = []
        for word in words:
            cleaned_word = replacements.get(word.upper(), word)
            cleaned_words.append(cleaned_word)

        return ' '.join(cleaned_words)

    def _clean_name(self, name: Any) -> str:
        """Clean and format names (contractor/business names)."""
        if name is None:
            return ""

        name = str(name).strip()

        # Title case for names, but preserve common business suffixes
        if name:
            # Handle common business suffixes
            suffixes = ['LLC', 'INC', 'CORP', 'LTD', 'CO', 'LP', 'LLP', 'PLLC']
            words = name.split()
            formatted_words = []

            for word in words:
                if word.upper() in suffixes:
                    formatted_words.append(word.upper())
                else:
                    formatted_words.append(word.title())

            return ' '.join(formatted_words)

        return name

    def _clean_address(self, address: Any) -> str:
        """Clean and format addresses."""
        if address is None:
            return ""

        address = str(address).strip()

        # Basic address formatting - remove extra whitespace and standardize
        address = re.sub(r'\s+', ' ', address)

        # Capitalize properly
        words = address.split()
        formatted_words = []

        street_suffixes = ['ST', 'AVE', 'BLVD', 'RD', 'LN', 'DR', 'CT', 'WAY', 'PL', 'CIR', 'PKWY', 'TRL', 'LOOP']
        directions = ['N', 'S', 'E', 'W', 'NE', 'NW', 'SE', 'SW', 'NORTH', 'SOUTH', 'EAST', 'WEST']

        for word in words:
            if word.upper() in street_suffixes or word.upper() in directions:
                formatted_words.append(word.upper())
            elif word.isdigit():
                formatted_words.append(word)
            else:
                formatted_words.append(word.title())

        return ' '.join(formatted_words)

    def _clean_description(self, description: Any) -> str:
        """Clean and format project descriptions."""
        if description is None:
            return ""

        description = str(description).strip()

        # Remove extra whitespace
        description = re.sub(r'\s+', ' ', description)

        # Capitalize first letter if not already capitalized
        if description and not description[0].isupper():
            description = description[0].upper() + description[1:]

        # Handle common abbreviations in descriptions
        description = description.replace(' EV ', ' Electric Vehicle ')
        description = description.replace(' EVSE', ' Electric Vehicle Supply Equipment')
        description = description.replace('CHG', 'Charging')

        return description

    def _format_date(self, date_str: Any) -> str:
        """Format dates in a user-friendly way."""
        if date_str is None:
            return ""

        date_str = str(date_str).strip()

        # If it's already in YYYY-MM-DD format, convert to MM/DD/YYYY
        if re.match(r'^\d{4}-\d{2}-\d{2}', date_str):
            try:
                parts = date_str[:10].split('-')
                year, month, day = parts
                return f"{month}/{day}/{year}"
            except:
                return date_str

        return date_str

    def _format_phone(self, phone: Any) -> str:
        """Format phone numbers in a readable format."""
        if phone is None:
            return ""

        phone = str(phone).strip()

        # Remove all non-digit characters
        digits = re.sub(r'\D', '', phone)

        # Format as (XXX) XXX-XXXX if we have 10 digits
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            # Handle numbers with country code
            return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
        elif digits:
            return digits  # Return digits if we can't format properly

        return ""

    def _truncate_text(self, text: str, max_length: int, add_ellipsis: bool = True) -> str:
        """
        Truncate text to optimal length for CSV column display.
        """
        if not text or len(text) <= max_length:
            return text

        if add_ellipsis and max_length > 3:
            return text[:max_length - 3] + "..."
        else:
            return text[:max_length]

    def _get_best_address(self, row: Dict[str, Any]) -> str:
        """
        Get the best available address from multiple possible fields.
        """
        address_fields = [
            "applicant_address",
            "contractor_address",
            "address",
            "property_address",
            "location_address",
            "site_address",
            "job_address",
            "project_address",
            "location"
        ]

        for field in address_fields:
            value = row.get(field)
            if value and str(value).strip():
                return self._clean_address(value)

        return "Address not available"

    def _get_best_contractor_name(self, row: Dict[str, Any]) -> str:
        """
        Get the best available contractor name from multiple possible fields.
        """
        name_fields = [
            "contractor_name",
            "applicant_name",
            "business_name",
            "company_name",
            "contractor_company_name",
            "applicant_company_name"
        ]

        for field in name_fields:
            value = row.get(field)
            if value and str(value).strip():
                return self._clean_name(value)

        return "Contractor not specified"

    def _get_best_phone(self, row: Dict[str, Any]) -> str:
        """
        Get the best available phone number from multiple possible fields.
        """
        phone_fields = [
            "contractor_phone",
            "applicant_phone",
            "phone",
            "contact_phone",
            "business_phone",
            "company_phone",
            "contractor_company_phone"
        ]

        for field in phone_fields:
            value = row.get(field)
            if value and str(value).strip():
                return self._format_phone(value)

        return "Phone not available"

    # ---------- Excel Export Methods ----------
    def create_excel_export(self, rows: List[Dict[str, Any]], filename: str = None) -> bytes:
        """
        Create Excel file with enhanced formatting, wider columns, and better text handling.
        Returns bytes that can be attached to emails or downloaded via web.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Permits Data"

        # Headers (no Relevance Score)
        headers = [
            "Project Scope",
            "Permit Type",
            "Date Issued",
            "Address",
            "Description",
            "Contractor Name",
            "Contact Phone",
            "Business Name"
        ]

        # Styling
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True, size=12)
        data_font = Font(size=11)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header row height
        ws.row_dimensions[1].height = 20

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = yellow_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Add data rows (no _rag_score handling)
        for row_idx, r in enumerate(rows or [], 2):
            ws.row_dimensions[row_idx].height = 20

            row_data = [
                self._clean_text(r.get("work_class", "")),
                self._clean_text(r.get("permit_type", "")),
                self._format_date(r.get("issued_date", "")),
                self._get_best_address(r),
                self._clean_description(r.get("description", "")),
                self._get_best_contractor_name(r),
                self._get_best_phone(r),
                self._clean_name(r.get("contractor_company_name", ""))
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col in (3, 7):  # Center align date + phone
                    cell.alignment = center_alignment
                else:
                    cell.alignment = left_alignment

        # Column widths
        column_widths = [23, 16, 12, 35, 120, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Metadata
        wb.properties.title = "Dumpster Rental Leads"
        wb.properties.creator = "Permits RAG System"
        wb.properties.description = f"Dumpster Rental Leads export of {len(rows)} permit records"

        # Save to bytes
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def save_excel_attachment(self, rows: List[Dict[str, Any]], filename: str = "Dumpster Rental Leads.xlsx",
                              include_score: bool = True) -> str:
        """
        Save Excel file to disk for attachment. Returns the full file path.
        """
        excel_data = self.create_excel_export(rows, include_score)

        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename = filename.rsplit('.', 1)[0] + '.xlsx'

        with open(filename, 'wb') as f:
            f.write(excel_data)

        return os.path.abspath(filename)

    def get_excel_for_download(self, rows: List[Dict[str, Any]], include_score: bool = True) -> tuple[bytes, str]:
        """
        Get Excel file data and suggested filename for web download.
        Returns (excel_bytes, suggested_filename)
        """
        import datetime

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Dumpster Rental Leads.xlsx"

        excel_data = self.create_excel_export(rows, include_score)

        return excel_data, filename

    def save_excel_file(self, rows: List[Dict[str, Any]], filename: str = "permits_data.xlsx",
                        include_score: bool = False) -> str:
        """
        Legacy method - use save_excel_attachment instead.

        Note: include_score parameter is kept for backward compatibility but ignored.
        """
        return self.save_excel_attachment(rows, filename)
        return self.save_excel_attachment(rows, filename)

    # ---------- Improved CSV Builders ----------
    def csv_from_rows(self, rows: List[Dict[str, Any]]) -> str:
        """
        Build a clean, user-friendly CSV from permit rows with optimized column widths.
        """
        out = io.StringIO()
        w = csv.writer(out)

        # User-friendly column headers
        w.writerow([
            "Project Scope",
            "Permit Type",
            "Date Issued",
            "Property Address",
            "Project Description",
            "Contractor Name",
            "Contact Phone",
            "Business Name"
        ])

        for r in rows:
            # Clean and format the data with optimal lengths using improved data extraction
            work_class = self._truncate_text(self._clean_text(r.get("work_class", "")), 15)
            permit_type = self._truncate_text(self._clean_text(r.get("permit_type", "")), 20)
            issued_date = self._format_date(r.get("issued_date", ""))
            address = self._truncate_text(self._get_best_address(r), 40)
            description = self._truncate_text(self._clean_description(r.get("description", "")), 60)
            contractor_name = self._truncate_text(self._get_best_contractor_name(r), 25)
            contractor_phone = self._get_best_phone(r)
            business_name = self._truncate_text(self._clean_name(r.get("contractor_company_name", "")), 25)

            w.writerow([
                work_class,
                permit_type,
                issued_date,
                address,
                description,
                contractor_name,
                contractor_phone,
                business_name
            ])
        return out.getvalue()

    def csv_full_from_rows(self, rows: List[Dict[str, Any]], include_score: bool = False) -> str:
        """
        Build a comprehensive CSV with clean formatting and wider column widths.
        Note: include_score parameter is kept for backward compatibility but ignored.
        """
        out = io.StringIO()
        w = csv.writer(out)

        # Headers (no Relevance Score)
        headers = [
            "Project Scope",
            "Permit Type",
            "Date Issued",
            "Address",
            "Description",
            "Contractor Name",
            "Contact Phone",
            "Business Name"
        ]

        w.writerow(headers)

        for r in rows or []:
            # Clean and format data with wider optimal lengths using improved data extraction
            row_data = [
                self._truncate_text(self._clean_text(r.get("work_class", "")), 25),
                self._truncate_text(self._clean_text(r.get("permit_type", "")), 30),
                self._format_date(r.get("issued_date", "")),
                self._truncate_text(self._get_best_address(r), 60),
                self._truncate_text(self._clean_description(r.get("description", "")), 150),
                self._truncate_text(self._get_best_contractor_name(r), 35),
                self._get_best_phone(r),
                self._truncate_text(self._clean_name(r.get("contractor_company_name", "")), 40)
            ]

            w.writerow(row_data)

        return out.getvalue()

    def _normalize_filters(self, filters: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
        """Convert string filters to list format for compatibility"""
        if not filters:
            return filters

        normalized = {}
        for key, value in filters.items():
            if isinstance(value, str) and value.strip():
                # Convert single string to list
                normalized[key] = [value.strip()]
            elif isinstance(value, list):
                # Keep lists as-is
                normalized[key] = value
            else:
                # Keep other types (dates, etc.)
                normalized[key] = value

        return normalized

    # ---------- Filters ----------
    def _apply_filters(self, rows: List[Dict[str, Any]], filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Supported filters (all optional).

        OR Logic Filters (default behavior):
          - city: List[str] - permit city must match ANY of the specified cities
          - permit_type: List[str] - permit type must match ANY of the specified types
          - permit_class_mapped: List[str] - permit class must match ANY of the specified classes
          - work_class: List[str] - work class must match ANY of the specified work classes
          - status: List[str] - status must match ANY of the specified statuses

        AND Logic Filters (strict matching):
          - city_strict_and: List[str] - permit city must contain ALL specified values
          - permit_type_strict_and: List[str] - permit type must contain ALL specified values
          - permit_class_mapped_strict_and: List[str] - permit class must contain ALL specified values
          - work_class_strict_and: List[str] - work class must contain ALL specified values
          - status_strict_and: List[str] - status must contain ALL specified values

        Date Filters:
          - issued_date_from: "YYYY-MM-DD"
          - issued_date_to: "YYYY-MM-DD"
          - applied_date_from: "YYYY-MM-DD"
          - applied_date_to: "YYYY-MM-DD"
        """
        if not filters:
            return rows

        def _in(val: Optional[str], allowed: Optional[List[str]]) -> bool:
            if not allowed:
                return True
            # AND logic: permit must match ALL specified values in the same field
            # This is useful for fields that can have multiple values or for strict matching
            val_str = str(val or "").lower().strip()
            allowed_lower = [str(a).lower().strip() for a in allowed]

            # If only one value is specified, use exact match
            if len(allowed_lower) == 1:
                return val_str == allowed_lower[0]

            # If multiple values are specified, permit must contain ALL of them
            # This is useful for multi-value fields or when you want strict matching
            return all(allowed_val in val_str for allowed_val in allowed_lower)

        def _in_strict_and(val: Optional[str], allowed: Optional[List[str]]) -> bool:
            """
            Strict AND logic: permit must match ALL specified values.
            Use this when you want permits that contain ALL the specified keywords/values.
            """
            if not allowed:
                return True
            val_str = str(val or "").lower().strip()
            allowed_lower = [str(a).lower().strip() for a in allowed]

            # Permit must contain ALL specified values
            return all(allowed_val in val_str for allowed_val in allowed_lower)

        def _date_ok(val: Optional[str], start: Optional[str], end: Optional[str]) -> bool:
            # If a range is specified and the row has no date, exclude it.
            if (start or end) and not val:
                return False
            if not val:
                return True
            try:
                d = val[:10]  # expect "YYYY-MM-DD..."
                if start and d < start:
                    return False
                if end and d > end:
                    return False
                return True
            except Exception:
                # if it's malformed, be conservative and exclude if a range is requested
                return not (start or end)

        cities = filters.get("city")
        ptypes = filters.get("permit_type")
        classes = filters.get("permit_class_mapped")
        wclasses = filters.get("work_class")
        statuses = filters.get("status")

        # New strict AND filters
        strict_cities = filters.get("city_strict_and")
        strict_ptypes = filters.get("permit_type_strict_and")
        strict_classes = filters.get("permit_class_mapped_strict_and")
        strict_wclasses = filters.get("work_class_strict_and")
        strict_statuses = filters.get("status_strict_and")

        issued_from = filters.get("issued_date_from")
        issued_to = filters.get("issued_date_to")
        applied_from = filters.get("applied_date_from")
        applied_to = filters.get("applied_date_to")

        out: List[Dict[str, Any]] = []
        for r in rows:
            # Apply regular OR filters
            if not _in(r.get("city"), cities):
                continue
            if not _in(r.get("permit_type"), ptypes):
                continue
            if not _in(r.get("permit_class_mapped"), classes):
                continue
            if not _in(r.get("work_class"), wclasses):
                continue
            if not _in(r.get("current_status"), statuses):
                continue

            # Apply strict AND filters (if specified)
            if strict_cities and not _in_strict_and(r.get("city"), strict_cities):
                continue
            if strict_ptypes and not _in_strict_and(r.get("permit_type"), strict_ptypes):
                continue
            if strict_classes and not _in_strict_and(r.get("permit_class_mapped"), strict_classes):
                continue
            if strict_wclasses and not _in_strict_and(r.get("work_class"), strict_wclasses):
                continue
            if strict_statuses and not _in_strict_and(r.get("current_status"), strict_statuses):
                continue

            if not _date_ok(r.get("issued_date"), issued_from, issued_to):
                continue
            if not _date_ok(r.get("applied_date"), applied_from, applied_to):
                continue

            out.append(r)
        return out

    # In your rag_engine_functional2.py, update _get_filtered_permits_from_db_simple method:

    def _get_filtered_permits_from_db_simple(self, filters: Dict[str, Any], limit: int) -> List[Dict[str, Any]]:
        """Robust database filter method - handles spacing and punctuation differences"""

        conn = self._connect()
        try:
            sql_parts = ["SELECT * FROM permits WHERE 1=1"]
            params = []

            # Helper function to normalize text for comparison
            def normalize_text(text):
                """Normalize text by removing extra spaces, standardizing punctuation"""
                if not text:
                    return ""
                # Convert to lowercase, replace multiple spaces with single space, strip
                normalized = ' '.join(str(text).lower().split())
                # Standardize common punctuation patterns
                normalized = normalized.replace(' - ', '-').replace('- ', '-').replace(' -', '-')
                normalized = normalized.replace(' & ', ' and ').replace('&', 'and')
                return normalized.strip()

            # Apply filters with robust matching
            if filters.get("city"):
                cities = filters["city"]
                if isinstance(cities, list) and cities:
                    city_conditions = []
                    for city in cities:
                        # Use simple case-insensitive matching for city (usually consistent)
                        city_conditions.append("LOWER(TRIM(city)) = LOWER(TRIM(?))")
                        params.append(city.strip())
                    sql_parts.append(f"AND ({' OR '.join(city_conditions)})")
                    logger.info(f"   📍 City filter applied: {cities}")

            if filters.get("permit_type"):
                permit_types = filters["permit_type"]
                if isinstance(permit_types, list) and permit_types:
                    type_conditions = []
                    for ptype in permit_types:
                        # Use simple case-insensitive matching for permit type (usually consistent)
                        type_conditions.append("LOWER(TRIM(permit_type)) = LOWER(TRIM(?))")
                        params.append(ptype.strip())
                    sql_parts.append(f"AND ({' OR '.join(type_conditions)})")
                    logger.info(f"   🏗 Permit type filter applied: {permit_types}")

            # ROBUST permit class matching - handles spacing differences
            if filters.get("permit_class_mapped"):
                permit_classes = filters["permit_class_mapped"]
                if isinstance(permit_classes, list) and permit_classes:
                    class_conditions = []
                    for pclass in permit_classes:
                        normalized_filter = normalize_text(pclass)
                        # Use REPLACE to normalize database values on the fly
                        class_conditions.append("""
                            REPLACE(REPLACE(LOWER(TRIM(permit_class_mapped)), ' - ', '-'), '- ', '-') 
                            = REPLACE(REPLACE(LOWER(TRIM(?)), ' - ', '-'), '- ', '-')
                        """.strip())
                        params.append(pclass.strip())
                        logger.info(f"         🏷 Matching '{pclass}' against normalized DB values")
                    sql_parts.append(f"AND ({' OR '.join(class_conditions)})")
                    logger.info(f"   🏷 Permit class filter applied: {permit_classes}")

            # ROBUST work class matching - handles case and spacing differences
            if filters.get("work_class"):
                work_classes = filters["work_class"]
                if isinstance(work_classes, list) and work_classes:
                    work_conditions = []
                    for work_class in work_classes:
                        normalized_filter = normalize_text(work_class)
                        # Use REPLACE to normalize database values and handle case differences
                        work_conditions.append("""
                            REPLACE(REPLACE(LOWER(TRIM(work_class)), ' and ', ' and '), 'and', 'and') 
                            = REPLACE(REPLACE(LOWER(TRIM(?)), ' and ', ' and '), 'and', 'and')
                        """.strip())
                        params.append(work_class.strip())
                        logger.info(f"         ⚒ Matching '{work_class}' against normalized DB values")
                    sql_parts.append(f"AND ({' OR '.join(work_conditions)})")
                    logger.info(f"   ⚒ Work class filter applied: {work_classes}")

            # Add limit and order
            sql_parts.append("ORDER BY issued_date DESC LIMIT ?")
            params.append(limit)

            # Execute query
            sql = " ".join(sql_parts)
            logger.info(f"   🗄 Final SQL: {sql}")

            cur = conn.cursor()
            cur.execute(sql, params)

            # Convert to dict
            columns = [desc[0] for desc in cur.description]
            results = [dict(zip(columns, row)) for row in cur.fetchall()]

            logger.info(f"   🗄 Database filter returned: {len(results)} permits")

            # DEBUG: If no results but filters seem reasonable, show what we're looking for
            if len(results) == 0 and filters:
                logger.warning(f"   🔍 DEBUG: No matches found. Let's check what we're looking for:")
                for key, values in filters.items():
                    if isinstance(values, list):
                        for value in values:
                            logger.warning(f"      {key}: '{value}' (normalized: '{normalize_text(value)}')")

            return results

        except Exception as e:
            logger.error(f"   ❌ Database filter error: {e}")
            return []
        finally:
            conn.close()

    def _search_within_filtered_permits(self, filtered_permits: List[Dict[str, Any]], query: str, top_k: int) -> List[
        Dict[str, Any]]:
        """Use FAISS to search within pre-filtered permits only"""

        if not filtered_permits or not query.strip():
            return filtered_permits[:top_k]

        # Get IDs of filtered permits
        filtered_ids = set(int(p['id']) for p in filtered_permits)

        # Create query embedding
        qvec = self._encode([query.strip()])[0].reshape(1, -1)

        # Search FAISS index - get more candidates than needed
        search_count = min(len(filtered_ids) * 2, 1000)
        sims, idxs = self.index.search(qvec, search_count)

        # Filter FAISS results to only include permits that passed database filters
        results = []
        for i, (idx, score) in enumerate(zip(idxs[0], sims[0])):
            if idx >= 0:  # Valid FAISS index
                permit_id = int(self.id_map[idx])
                if permit_id in filtered_ids:
                    # Find the full permit data from our filtered permits
                    permit_data = next((p for p in filtered_permits if int(p['id']) == permit_id), None)
                    if permit_data:
                        permit_data['_rag_score'] = float(score)
                        results.append(permit_data)

                        if len(results) >= top_k:
                            break

        logger.info(f"   🎯 FAISS within filtered: {len(results)} semantic matches")
        return results

    def _get_recent_permits_simple(self, limit: int) -> List[Dict[str, Any]]:
        """Get recent permits without any filters"""

        conn = self._connect()
        try:
            cur = conn.cursor()
            cur.execute("SELECT * FROM permits ORDER BY issued_date DESC LIMIT ?", (limit,))

            columns = [desc[0] for desc in cur.description]
            results = [dict(zip(columns, row)) for row in cur.fetchall()]

            logger.info(f"   🗄️ Recent permits returned: {len(results)} permits")
            return results

        except Exception as e:
            logger.error(f"   ❌ Recent permits error: {e}")
            return []
        finally:
            conn.close()

    # In your rag_engine_functional2.py, replace _semantic_search_within_permits method:

    def _semantic_search_within_permits(self, permits: List[Dict[str, Any]], query: str, top_k: int,
                                        return_scores: bool) -> List[Dict[str, Any]]:
        """Use FAISS to search within specific permits - FIXED VERSION"""

        if not permits:
            logger.info(f"   ⚠️ No permits to search within")
            return []

        try:
            logger.info(f"   🔍 SEMANTIC SEARCH DEBUG:")
            logger.info(f"      📊 Input permits: {len(permits)}")
            logger.info(f"      🔎 Query: '{query}'")

            # Check if index is loaded
            if self.index is None or self.id_map is None:
                logger.warning(f"      ❌ FAISS index not loaded, attempting to load...")
                if not self.load():
                    logger.error(f"      ❌ Failed to load FAISS index")
                    return self._simple_text_search(permits, query, top_k, return_scores)
                logger.info(f"      ✅ FAISS index loaded: {len(self.id_map)} vectors")

            # Get permit IDs from filtered results
            permit_ids = set(int(p['id']) for p in permits)
            logger.info(f"      🆔 Target permit IDs: {sorted(list(permit_ids))[:10]}...")  # Show first 10

            # DEBUG: Check what IDs are in the FAISS index
            faiss_ids = set(self.id_map.tolist())
            logger.info(f"      🗂️ FAISS index contains {len(faiss_ids)} IDs")
            logger.info(f"      🗂️ Sample FAISS IDs: {sorted(list(faiss_ids))[:10]}...")

            # Check overlap between filtered permits and FAISS index
            overlap = permit_ids.intersection(faiss_ids)
            logger.info(f"      🔗 Overlap: {len(overlap)} permits exist in both filtered results and FAISS index")

            if len(overlap) == 0:
                logger.warning(f"      ⚠️ No overlap between filtered permits and FAISS index!")
                logger.warning(f"      ⚠️ This suggests the FAISS index is out of sync with database")
                # Fallback to simple text search
                return self._simple_text_search(permits, query, top_k, return_scores)

            # Create query embedding
            query_embedding = self._encode([query.strip()])[0].reshape(1, -1)
            logger.info(f"      🧮 Query embedding shape: {query_embedding.shape}")

            # Search FAISS - get more results than needed
            search_count = min(len(self.id_map), top_k * 5)
            scores, indices = self.index.search(query_embedding, search_count)
            logger.info(f"      🔍 FAISS search returned {len(indices[0])} results")

            # Match FAISS results with filtered permits
            results = []
            for idx, score in zip(indices[0], scores[0]):
                if idx >= 0:  # Valid FAISS index
                    permit_id = int(self.id_map[idx])

                    # Only include if this permit was in our filtered results
                    if permit_id in permit_ids:
                        # Find the full permit data from filtered permits
                        permit_data = next((p for p in permits if int(p['id']) == permit_id), None)
                        if permit_data:
                            if return_scores:
                                permit_data['_rag_score'] = float(score)
                            results.append(permit_data)

                            logger.info(f"         ✅ Match: Permit {permit_id}, Score: {score:.3f}")

                            if len(results) >= top_k:
                                break

            logger.info(f"      🎯 Final semantic results: {len(results)} permits")

            if len(results) == 0:
                logger.warning(f"      ⚠️ Semantic search found no matches, falling back to text search")
                return self._simple_text_search(permits, query, top_k, return_scores)

            return results

        except Exception as e:
            logger.error(f"      ❌ Semantic search error: {e}")
            logger.error(f"      ↳ Falling back to simple text search")
            return self._simple_text_search(permits, query, top_k, return_scores)

    def _simple_text_search(self, permits: List[Dict[str, Any]], query: str, top_k: int, return_scores: bool) -> List[
        Dict[str, Any]]:
        """Fallback text search in descriptions"""

        query_lower = query.lower().strip()
        scored_permits = []

        for permit in permits:
            description = str(permit.get('description', '')).lower()
            if query_lower in description:
                score = description.count(query_lower) * 10
                if return_scores:
                    permit['_rag_score'] = score
                scored_permits.append(permit)

        # Sort by score if available
        if return_scores:
            scored_permits.sort(key=lambda x: x.get('_rag_score', 0), reverse=True)

        return scored_permits[:top_k]

    # ---------- Search ----------
    # Fix 2: Replace your current search() method with this filter-first approach
    # In your rag_engine_functional2.py, update search_fixed method to remove the debug bypass:

    def search_fixed(self, query: str, top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
                     oversample: int = 5, max_oversample: int = 80, return_scores: bool = False) -> List[
        Dict[str, Any]]:
        res = self.unified_search(query=query, mode="semantic", filters=filters, top_k=top_k,
                                  oversample=oversample, return_scores=return_scores, debug=False)
        return res.get("results", [])



    def search_description_only(self, query: str, top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
                                oversample: int = 5, max_oversample: int = 80, return_scores: bool = False) -> List[
        Dict[str, Any]]:
        """Same as search() - filter-first approach"""
        return self.search(query, top_k, filters, oversample, max_oversample, return_scores)

    def search_keywords_in_description(
            self,
            keywords: str,
            top_k: int = 20,
            filters: Optional[Dict[str, Any]] = None,
            return_scores: bool = False,
    ) -> List[Dict[str, Any]]:
        """
        Simple keyword search in permit descriptions using SQL LIKE.
        This bypasses FAISS and directly searches the database for exact keyword matches.

        Args:
            keywords: Keywords to search for (e.g., "roof", "electrical")
            top_k: Maximum number of results
            filters: Optional filters
            return_scores: Not applicable for keyword search (always False)

        Returns:
            List of permit rows with descriptions containing the keywords
        """
        conn = self._connect()
        try:
            # Build the SQL query
            sql_parts = ["SELECT * FROM permits WHERE 1=1"]
            params = []

            # Add keyword search in description
            if keywords and keywords.strip():
                sql_parts.append("AND description LIKE ?")
                params.append(f"%{keywords.strip()}%")

            # Add filters
            if filters:
                if filters.get("city"):
                    cities = filters["city"]
                    if isinstance(cities, list) and cities:
                        placeholders = ",".join("?" * len(cities))
                        sql_parts.append(f"AND city IN ({placeholders})")
                        params.extend(cities)

                if filters.get("permit_type"):
                    ptypes = filters["permit_type"]
                    if isinstance(ptypes, list) and ptypes:
                        placeholders = ",".join("?" * len(ptypes))
                        sql_parts.append(f"AND permit_type IN ({placeholders})")
                        params.extend(ptypes)

                if filters.get("work_class"):
                    wclasses = filters["work_class"]
                    if isinstance(wclasses, list) and wclasses:
                        placeholders = ",".join("?" * len(wclasses))
                        sql_parts.append(f"AND work_class IN ({placeholders})")
                        params.extend(wclasses)

                if filters.get("issued_date_from"):
                    sql_parts.append("AND issued_date >= ?")
                    params.append(filters["issued_date_from"])

                if filters.get("issued_date_to"):
                    sql_parts.append("AND issued_date <= ?")
                    params.append(filters["issued_date_to"])

            # Add limit and order
            sql_parts.append("ORDER BY issued_date DESC LIMIT ?")
            params.append(top_k)

            # Execute query
            sql = " ".join(sql_parts)
            cur = conn.cursor()
            cur.execute(sql, params)

            # Convert to list of dicts
            cols = [desc[0] for desc in cur.description]
            rows = []
            for row in cur.fetchall():
                row_dict = dict(zip(cols, row))
                if return_scores:
                    row_dict["_rag_score"] = 1.0  # Keyword matches get score 1.0
                rows.append(row_dict)

            return rows

        finally:
            conn.close()

    # ---------- 75/25 Distribution Methods ----------
    def distribute_permits_with_ratio(self, permits: List[Dict[str, Any]], client_count: int,
                                      ratio: List[float] = None) -> List[List[Dict[str, Any]]]:
        """
        Distribute permits among clients based on a given ratio.

        Args:
            permits: List of permit dictionaries
            client_count: Number of clients to distribute among
            ratio: List of ratios (e.g., [0.75, 0.25] for 75/25 split)

        Returns:
            List of permit lists, one for each client
        """
        if not permits or client_count <= 0:
            return [[] for _ in range(client_count)]

        # Validate ratio
        if ratio is None:
            # Default to equal distribution
            ratio = [1.0 / client_count] * client_count
        elif len(ratio) != client_count:
            raise ValueError(f"Ratio length ({len(ratio)}) must match client count ({client_count})")

        # Normalize ratio to sum to 1.0
        total_ratio = sum(ratio)
        if total_ratio == 0:
            ratio = [1.0 / client_count] * client_count
        else:
            ratio = [r / total_ratio for r in ratio]

        # Calculate permit counts for each client
        total_permits = len(permits)
        client_counts = []
        remaining_permits = total_permits

        for i, r in enumerate(ratio):
            if i == len(ratio) - 1:
                # Last client gets all remaining permits
                count = remaining_permits
            else:
                count = int(round(total_permits * r))
                remaining_permits -= count
            client_counts.append(count)

        # Distribute permits
        result = []
        start_idx = 0

        for count in client_counts:
            end_idx = start_idx + count
            result.append(permits[start_idx:end_idx])
            start_idx = end_idx

        return result

    # --- Add near the other search methods in RAGIndex class ---
    def unified_search(
            self,
            query: str,
            mode: str = "dual",
            filters: Optional[Dict[str, Any]] = None,
            top_k: int = 20,
            oversample: int = 5,
            return_scores: bool = False,
            debug: bool = False,
    ) -> Dict[str, Any]:
        """
        Unified search entry point.
        mode: "keyword" | "semantic" | "dual"
        Returns:
          - if mode == "keyword" -> {"mode":"keyword","results":[...],"count":N}
          - if mode == "semantic" -> {"mode":"semantic","results":[...],"count":N}
          - if mode == "dual" -> {"mode":"dual","results":{"keyword":[...],"semantic":[...]}, "counts":{...}}
        """
        lg = logger.debug if debug else logger.info
        lg(f"🔎 unified_search(mode={mode}, top_k={top_k}, oversample={oversample})")
        lg(f"   query='{query}'")
        lg(f"   filters={filters}")

        filters = self._normalize_filters(filters or {})

        # Prepare filtered permits for semantic branch
        filtered_permits: List[Dict[str, Any]] = []
        if mode in ("semantic", "dual"):
            if filters and any(filters.values()):
                db_limit = max(top_k * oversample, 1000)
                filtered_permits = self._get_filtered_permits_from_db_simple(filters, db_limit)
                lg(f"   📊 DB-filtered for semantic: {len(filtered_permits)} (limit={db_limit})")
            else:
                db_limit = max(top_k * 3, 500)
                filtered_permits = self._get_recent_permits_simple(db_limit)
                lg(f"   📊 Recent (no filters) for semantic: {len(filtered_permits)} (limit={db_limit})")

        # KEYWORD branch
        if mode == "keyword":
            kw = self.search_keywords_in_description(
                keywords=query, top_k=top_k, filters=filters, return_scores=return_scores
            )
            return {"mode": "keyword", "results": kw, "count": len(kw)}

        # SEMANTIC branch
        if mode == "semantic":
            if not filtered_permits:
                return {"mode": "semantic", "results": [], "count": 0}
            if (self.index is None or self.id_map is None) and not self.load():
                lg("   ⚠️ FAISS not loaded, falling back to simple text search within filtered permits")
                results = self._simple_text_search(filtered_permits, query, top_k, return_scores)
                return {"mode": "semantic", "results": results, "count": len(results)}
            sem = self._semantic_search_within_permits(filtered_permits, query, top_k, return_scores)
            return {"mode": "semantic", "results": sem, "count": len(sem)}

        # DUAL branch
        # Keyword side: use keyword SQL (no score or optional)
        kw = self.search_keywords_in_description(
            keywords=query, top_k=top_k, filters=filters, return_scores=False
        )
        # Semantic side
        if filtered_permits:
            if (self.index is None or self.id_map is None) and not self.load():
                lg("   ⚠️ FAISS not loaded, falling back to text search for semantic half")
                sem = self._simple_text_search(filtered_permits, query, top_k, return_scores)
            else:
                sem = self._semantic_search_within_permits(filtered_permits, query, top_k, return_scores)
        else:
            sem = []

        return {
            "mode": "dual",
            "results": {"keyword": kw, "semantic": sem},
            "counts": {"keyword": len(kw), "semantic": len(sem)}
        }

    def search(self, query: str, top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
               oversample: int = 5, max_oversample: int = 80, return_scores: bool = False) -> List[Dict[str, Any]]:
        """
        Compatibility wrapper - calls search_fixed()
        This ensures backward compatibility if any code calls .search()
        """
        return self.search_fixed(
            query=query,
            top_k=top_k,
            filters=filters,
            oversample=oversample,
            max_oversample=max_oversample,
            return_scores=return_scores
        )

    def search_and_distribute(self, query: str, client_count: int, ratio: List[float] = None,
                              top_k: int = 20, filters: Optional[Dict[str, Any]] = None,
                              oversample: int = 5, return_scores: bool = True) -> List[List[Dict[str, Any]]]:
        """
        Search for permits and distribute them among clients based on ratio.

        Args:
            query: Search query
            client_count: Number of clients
            ratio: Distribution ratio (e.g., [0.75, 0.25])
            top_k: Number of permits to retrieve
            filters: Optional filters
            oversample: Oversampling factor
            return_scores: Whether to include relevance scores

        Returns:
            List of permit lists, one for each client
        """
        # Search for permits
        permits = self.search_fixed(
            query=query,
            top_k=top_k,
            filters=filters,
            oversample=oversample,
            return_scores=return_scores
        )

        # Distribute based on ratio
        return self.distribute_permits_with_ratio(permits, client_count, ratio)

    def load_with_debug(self) -> bool:
        """Enhanced load method with debugging"""
        print("there")
        print(f"🔄 LOADING INDEX FROM DISK:")
        print(f"   📁 Index dir: {self.index_dir}")
        print(f"   📄 Index path: {self.index_path}")
        print(f"   📄 ID map path: {self.idmap_path}")

        try:
            # Check if files exist
            if not os.path.exists(self.index_path):
                print(f"   ❌ FAISS index file not found: {self.index_path}")
                return False

            if not os.path.exists(self.idmap_path):
                print(f"   ❌ ID map file not found: {self.idmap_path}")
                return False

            # Load FAISS index
            print(f"   📈 Loading FAISS index...")
            import faiss
            self.index = faiss.read_index(self.index_path)
            print(f"   ✅ FAISS index loaded: {self.index.ntotal} vectors")

            # Load ID mapping
            print(f"   🗂️ Loading ID mapping...")
            self.id_map = np.load(self.idmap_path)
            print(f"   ✅ ID map loaded: {len(self.id_map)} entries")

            # Verify consistency
            if self.index.ntotal != len(self.id_map):
                print(f"   ⚠️ WARNING: FAISS vectors ({self.index.ntotal}) != ID map entries ({len(self.id_map)})")
                return False

            print(f"   ✅ Index successfully loaded into memory")
            return True

        except Exception as e:
            print(f"   ❌ Error loading index: {e}")
            import traceback
            traceback.print_exc()
            return False

    # Add this method to your rag_engine_functional.py file

    def search_description_only_with_ids(self, query: str, permit_ids: List[int], top_k: int = 20) -> List[
        Dict[str, Any]]:
        """
        Search ONLY in descriptions of specific permit IDs
        This ensures we only search within pre-filtered results
        """
        import sqlite3

        if not permit_ids:
            return []

        try:
            conn = sqlite3.connect(self.db_path)
            cur = conn.cursor()

            # Get only the specified permits
            placeholders = ",".join("?" for _ in permit_ids)
            query_sql = f"""
                SELECT * FROM permits 
                WHERE id IN ({placeholders})
                AND description IS NOT NULL 
                AND description != ''
                AND description LIKE ?
                ORDER BY 
                    CASE 
                        WHEN description LIKE ? THEN 1
                        WHEN description LIKE ? THEN 2  
                        ELSE 3
                    END,
                    issued_date DESC
                LIMIT ?
            """

            # Create search patterns
            query_lower = query.lower()
            exact_pattern = f"%{query_lower}%"
            word_pattern = f"% {query_lower} %"
            start_pattern = f"{query_lower}%"

            params = permit_ids + [exact_pattern, word_pattern, start_pattern, top_k]

            cur.execute(query_sql, params)
            rows = cur.fetchall()

            # Convert to dict format
            columns = [description[0] for description in cur.description]
            results = []

            for i, row in enumerate(rows):
                row_dict = dict(zip(columns, row))
                # Add simple relevance score
                description = row_dict.get("description", "").lower()
                score = description.count(query_lower) * 10  # Simple frequency-based scoring
                if query_lower in description.split():  # Bonus for exact word match
                    score += 5
                row_dict["_rag_score"] = score
                results.append(row_dict)

            # Sort by score descending
            results.sort(key=lambda x: x.get("_rag_score", 0), reverse=True)

            conn.close()
            return results

        except Exception as e:
            logger.error(f"Error in search_description_only_with_ids: {e}")
            return []

    def search_and_distribute_75_25(self, query: str, top_k: int = 20,
                                    filters: Optional[Dict[str, Any]] = None,
                                    oversample: int = 5, return_scores: bool = True) -> Tuple[
        List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Convenience method for 75/25 distribution between exactly 2 clients.

        Args:
            query: Search query
            top_k: Number of permits to retrieve
            filters: Optional filters
            oversample: Oversampling factor
            return_scores: Whether to include relevance scores

        Returns:
            Tuple of (client1_permits, client2_permits)
        """
        distributed = self.search_and_distribute(
            query=query,
            client_count=2,
            ratio=[0.75, 0.25],
            top_k=top_k,
            filters=filters,
            oversample=oversample,
            return_scores=return_scores
        )

        if len(distributed) != 2:
            raise RuntimeError("Expected exactly 2 client distributions")

        return distributed[0], distributed[1]

    # ---------- Convenience ----------
    def search_by_filters(
            self,
            filters: Optional[Dict[str, Any]] = None,
            top_k: int = 100,
    ) -> List[Dict[str, Any]]:
        """Filter-only retrieval (alias for search(query="", filters=...))."""
        return self.search(query="", top_k=top_k, filters=filters, oversample=1)


_all_ = ["RAGIndex", "_row_to_text", "_safe"]
